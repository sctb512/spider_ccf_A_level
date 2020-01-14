from urllib.request import urlopen
from bs4 import BeautifulSoup
import re
import openpyxl


def getBsObj(pageUrl: str):
    try:
        html = urlopen(pageUrl)
    except:
        return None
    else:
        return BeautifulSoup(html)

def getVolume(url):
    bsObj = getBsObj(url)
    list = bsObj.find_all('a',text=re.compile("Volume.*"))
    hrefs=[]
    for i in list:
        hrefs.append(i.get('href'))
    return hrefs

def init(path):
    book = openpyxl.Workbook()
    book.save(path)


def getInfo(url, sname, name, type):
    bsObj=getBsObj(url)

    wb = openpyxl.load_workbook("data.xlsx")
    ws = wb[type]

    for i in bsObj.find_all(class_="entry article"):
        title=i.find(class_="title")
        title=title.get_text()#提取文本
        doi=i.find("a")#找到规律，直接拿第一个链接就是了,doi值隐藏在这个链接之后
        doi=doi.get("href")
        doi=str(doi)
        doi=doi.replace("http://","")#除去链接头
        doi=doi.replace("https://","")
        ws.append([title, doi, sname, name])

    wb.save("data.xlsx")



if __name__ == '__main__':

    init("data.xlsx")
    mainPage = "http://www.ccf.org.cn"
    bsObj = getBsObj(mainPage + "/xspj/gyml/")
    repeatUrl = set()  # 用于去掉重复的url爬取
    repeatUrl.add(mainPage + "/xspj/gyml/")
    for link in bsObj.find("div", {"class": "col-md-2"}).find_all("a"):
        href = mainPage + (link.attrs['href'])  # 拿到a标签里面的href属性值
        if href not in repeatUrl:
            print("#爬取分类目录的A类论文")
            repeatUrl.add(href)
            print("bsoj href:"+href)
            bsObj = getBsObj(href)

            try:  # 加上trycatch语句，因为可以看到目录的最后一个是无关的“联系我们”
                typeObj = bsObj.find("div", class_="m-text-mg").find_all("h4")
                type = typeObj[1].text.replace("(", "").replace(")", "").replace("/","_").replace(" ","")
                print(type)
                wb = openpyxl.load_workbook("data.xlsx")
                wb.create_sheet(type)
                ws = wb[type]
                ws.append(["title", "doi", "sname", "name"])
                wb.save("data.xlsx")


                for li in bsObj.find("ul", class_="g-ul x-list3").find_all("li")[1:]:

                    sname = li.find("div", class_="sname").text
                    name = li.find_all("div")[2].text
                    href = li.find("a").attrs['href']

                    if href in repeatUrl:
                        continue

                    print("#爬取volume:"+href)
                    repeatUrl.add(href)
                    list = getVolume(href)
                    for a in list:
                        if a in repeatUrl:
                            continue
                        print("#爬取文章:"+a)
                        repeatUrl.add(a)
                        getInfo(a, sname, name, type)

            except Exception as e:  # 没有相关内容的跳过就好
                print(e)